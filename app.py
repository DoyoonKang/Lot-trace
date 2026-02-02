# app.py
import streamlit as st
import pandas as pd
import altair as alt
import datetime as dt
import re
from pathlib import Path
from io import StringIO
import requests
from openpyxl import load_workbook

# ==========================================================
# Page / Style
# ==========================================================
st.set_page_config(page_title="ì•¡ìƒ ì‰í¬ Lot ì¶”ì  ê´€ë¦¬", page_icon="ğŸ§ª", layout="wide")

st.markdown(
    """
    <style>
      /* ì „ì²´ ë ˆì´ì•„ì›ƒ */
      .block-container { 
        padding-top: 1.5rem; 
        padding-bottom: 2rem; 
        max-width: 1400px;
      }
      
      /* íƒ€ì´í‹€ */
      h1 {
        color: #1f2937;
        font-weight: 800;
        font-size: 2.2rem !important;
        margin-bottom: 0.5rem !important;
      }
      
      /* KPI ì¹´ë“œ ìŠ¤íƒ€ì¼ */
      div[data-testid="metric-container"] {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        border-radius: 12px;
        padding: 1rem;
        box-shadow: 0 4px 6px rgba(0,0,0,0.1);
      }
      
      div[data-testid="metric-container"] label {
        color: white !important;
        font-weight: 700 !important;
        font-size: 0.85rem !important;
      }
      
      div[data-testid="metric-container"] [data-testid="stMetricValue"] {
        color: white !important;
        font-size: 1.8rem !important;
        font-weight: 900 !important;
      }
      
      /* ì„¹ì…˜ íƒ€ì´í‹€ */
      .section-title { 
        font-size: 1.4rem; 
        font-weight: 900; 
        margin: 1.5rem 0 0.5rem 0;
        color: #1f2937;
        border-left: 5px solid #667eea;
        padding-left: 12px;
      }
      
      .section-sub { 
        color: #6b7280; 
        font-size: 0.95rem; 
        margin-bottom: 1rem;
        padding-left: 17px;
      }
      
      /* KPI ë…¸íŠ¸ */
      .kpi-note { 
        color: #6b7280; 
        font-size: 0.88rem; 
        margin-top: 0.5rem;
        padding: 0.8rem;
        background: #f3f4f6;
        border-radius: 8px;
        border-left: 3px solid #fbbf24;
      }
      
      /* Expander ìŠ¤íƒ€ì¼ */
      div[data-testid="stExpander"] {
        border: 2px solid #e5e7eb;
        border-radius: 10px;
        background: #ffffff;
      }
      
      div[data-testid="stExpander"] > details > summary { 
        font-weight: 800;
        color: #374151;
        font-size: 1.05rem;
      }
      
      /* íƒ­ ìŠ¤íƒ€ì¼ */
      .stTabs [data-baseweb="tab-list"] {
        gap: 8px;
        background-color: #f9fafb;
        padding: 8px;
        border-radius: 10px;
      }
      
      .stTabs [data-baseweb="tab"] {
        border-radius: 8px;
        padding: 10px 20px;
        font-weight: 700;
        font-size: 0.95rem;
      }
      
      .stTabs [aria-selected="true"] {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white !important;
      }
      
      /* í…Œì´ë¸” ìŠ¤íƒ€ì¼ */
      .dataframe {
        border-radius: 10px !important;
        overflow: hidden;
        box-shadow: 0 2px 4px rgba(0,0,0,0.05);
      }
      
      /* ë²„íŠ¼ ìŠ¤íƒ€ì¼ */
      .stButton > button {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
        border: none;
        border-radius: 8px;
        padding: 0.6rem 1.5rem;
        font-weight: 700;
        transition: all 0.3s;
      }
      
      .stButton > button:hover {
        transform: translateY(-2px);
        box-shadow: 0 4px 12px rgba(102, 126, 234, 0.4);
      }
      
      /* ë‹¤ìš´ë¡œë“œ ë²„íŠ¼ */
      .stDownloadButton > button {
        background: linear-gradient(135deg, #10b981 0%, #059669 100%);
        color: white;
        border: none;
        border-radius: 8px;
        font-weight: 700;
      }
      
      /* ê²½ê³  ë°•ìŠ¤ */
      .stAlert {
        border-radius: 10px;
        border-left: 5px solid;
      }
      
      /* ì°¨íŠ¸ ì»¨í…Œì´ë„ˆ */
      .chart-container {
        background: white;
        padding: 1.5rem;
        border-radius: 12px;
        box-shadow: 0 2px 8px rgba(0,0,0,0.08);
        margin-bottom: 1rem;
      }
      
      /* í¼ ìŠ¤íƒ€ì¼ */
      .stForm {
        background: #f9fafb;
        padding: 1.5rem;
        border-radius: 12px;
        border: 2px solid #e5e7eb;
      }
      
      /* ì…ë ¥ í•„ë“œ */
      .stTextInput > div > div > input,
      .stNumberInput > div > div > input,
      .stSelectbox > div > div > div {
        border-radius: 8px;
        border: 2px solid #e5e7eb;
      }
      
      /* êµ¬ë¶„ì„  */
      hr {
        margin: 2rem 0;
        border: none;
        border-top: 2px solid #e5e7eb;
      }
      
      /* ì‚¬ì´ë“œë°” */
      .css-1d391kg, [data-testid="stSidebar"] {
        background: linear-gradient(180deg, #f9fafb 0%, #ffffff 100%);
      }
      
      /* ìƒíƒœ ë°°ì§€ */
      .status-badge-success {
        background: #10b981;
        color: white;
        padding: 0.3rem 0.8rem;
        border-radius: 20px;
        font-size: 0.85rem;
        font-weight: 700;
        display: inline-block;
      }
      
      .status-badge-warning {
        background: #f59e0b;
        color: white;
        padding: 0.3rem 0.8rem;
        border-radius: 20px;
        font-size: 0.85rem;
        font-weight: 700;
        display: inline-block;
      }
      
      .status-badge-error {
        background: #ef4444;
        color: white;
        padding: 0.3rem 0.8rem;
        border-radius: 20px;
        font-size: 0.85rem;
        font-weight: 700;
        display: inline-block;
      }
    </style>
    """,
    unsafe_allow_html=True,
)

# toast
if "_toast_msg" not in st.session_state:
    st.session_state["_toast_msg"] = None
if st.session_state.get("_toast_msg"):
    try:
        st.toast(st.session_state["_toast_msg"])
    except Exception:
        pass
    st.session_state["_toast_msg"] = None

# ==========================================================
# Config
# ==========================================================
DEFAULT_XLSX = "ì•¡ìƒì‰í¬_Lotì¶”ì ê´€ë¦¬_FINAL.xlsx"
DEFAULT_STOCK_XLSX = "ì•¡ìƒ ì¬ê³ ì¡°ì‚¬í‘œ_ìë™ê³„ì‚° (12).xlsx"

SHEET_BINDER = "ë°”ì¸ë”_ì œì¡°_ì…ê³ "
SHEET_SINGLE = "ë‹¨ì¼ìƒ‰_ìˆ˜ì…ê²€ì‚¬"
SHEET_SPEC_BINDER = "Spec_Binder"
SHEET_SPEC_SINGLE = "Spec_Single_H&S"
SHEET_BASE_LAB = "ê¸°ì¤€LAB"
SHEET_BINDER_RETURN = "ë°”ì¸ë”_ì—…ì²´ë°˜í™˜"  # ì—†ìœ¼ë©´ ìë™ ìƒì„±

# ë°”ì¸ë” ì…ì¶œê³ (êµ¬ê¸€ì‹œíŠ¸)
BINDER_SHEET_ID = "1H2fFxnf5AvpSlu-uoZ4NpTv8LYLNwTNAzvlntRQ7FS8"
BINDER_SHEET_HEMA = "HEMA ë°”ì¸ë” ì…ì¶œê³  ê´€ë¦¬ëŒ€ì¥"
BINDER_SHEET_SIL = "Siliconë°”ì¸ë” ì…ì¶œê³  ê´€ë¦¬ëŒ€ì¥"

COLOR_KEYS = ["BLACK", "BLUE", "GREEN", "YELLOW", "RED", "PINK", "WHITE", "OTHER"]

# ==========================================================
# Helpers
# ==========================================================
def norm_key(x) -> str:
    if x is None:
        return ""
    s = str(x).replace("\n", " ").replace("\r", " ").strip()
    s = re.sub(r"\s+", " ", s)
    return s

def find_col(df: pd.DataFrame, want: str):
    w = norm_key(want)
    for c in df.columns:
        if norm_key(c) == w:
            return c
    return None

def normalize_date(x):
    if x is None or (isinstance(x, float) and pd.isna(x)):
        return None
    if isinstance(x, (dt.date, dt.datetime)):
        return x.date() if isinstance(x, dt.datetime) else x
    try:
        return pd.to_datetime(x).date()
    except Exception:
        return None

def safe_date_bounds(series: pd.Series):
    s = pd.to_datetime(series, errors="coerce").dropna()
    if len(s) == 0:
        today = dt.date.today()
        return today, today
    return s.min().date(), s.max().date()

def detect_date_col(df: pd.DataFrame):
    for c in df.columns:
        ck = norm_key(c).lower()
        if any(k in ck for k in ["ì¼ì", "ë‚ ì§œ", "date", "ì…ê³ ì¼", "ì¶œê³ ì¼", "ë°˜ì…ì¼", "ë°˜ì¶œì¼"]):
            return c
    return None

def file_sig(path: str):
    """ìºì‹œ ë¬´íš¨í™”ë¥¼ ìœ„í•œ ì‹œê·¸ë‹ˆì²˜"""
    try:
        p = Path(path)
        if not p.exists():
            return None
        stat = p.stat()
        return (str(p.resolve()), int(stat.st_size), int(stat.st_mtime))
    except Exception:
        return None

def ensure_sheet_exists(xlsx_path: str, sheet_name: str, headers: list[str]):
    wb = load_workbook(xlsx_path)
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
        ws.append(headers)
        wb.save(xlsx_path)

def add_excel_row_number(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["_excel_row"] = df.index + 2  # í—¤ë” 1í–‰ ê°€ì •
    return df

# ==========================================================
# Color helpers (ìš”ì²­: BLACK/RED ë“± ëŒ€ë¬¸ì ê°€ì‹œí™”)
# ==========================================================
def normalize_color_group(x) -> str:
    if x is None or (isinstance(x, float) and pd.isna(x)):
        return "OTHER"
    s = str(x).strip()
    if not s or s.lower() in ("nan", "none"):
        return "OTHER"

    u = s.upper()
    if "BLACK" in u or "ê²€ì •" in s or "í‘" in s:
        return "BLACK"
    if "WHITE" in u or "í°" in s or "ë°±" in s:
        return "WHITE"
    if "RED" in u or "ë¹¨" in s or "ì " in s:
        return "RED"
    if "YELLOW" in u or "ë…¸" in s or "í™©" in s or "ì˜" in s:
        return "YELLOW"
    if "GREEN" in u or "ì´ˆ" in s or "ë…¹" in s:
        return "GREEN"
    if "BLUE" in u or "íŒŒ" in s or "ì²­" in s:
        return "BLUE"
    if "PINK" in u or "í•‘" in s:
        return "PINK"

    if u in COLOR_KEYS:
        return u
    return "OTHER"

def normalize_product_code(x) -> str:
    if x is None or (isinstance(x, float) and pd.isna(x)):
        return ""
    s = str(x).strip()
    if not s or s.lower() in ("nan", "none"):
        return ""
    s = s.replace("â€“", "-").replace("â€”", "-").replace("âˆ’", "-")
    s = re.sub(r"\s+", " ", s).strip()
    s = s.replace("(ì•¡ìƒì‰í¬)", "").replace("ì•¡ìƒì‰í¬", "").strip()
    return s

def _color_scale_color_group():
    domain = ["BLACK", "BLUE", "GREEN", "YELLOW", "RED", "PINK", "WHITE", "OTHER"]
    rng = ["#111111", "#1f77b4", "#2ca02c", "#f1c40f", "#d62728", "#e377c2", "#dddddd", "#7f7f7f"]
    return alt.Scale(domain=domain, range=rng)

# ==========================================================
# Product -> ColorGroup mapping (Spec + Single)
# ==========================================================
def build_product_to_color_map(spec_single: pd.DataFrame, single_df: pd.DataFrame) -> dict[str, str]:
    mapping: dict[str, str] = {}

    sp_pc = find_col(spec_single, "ì œí’ˆì½”ë“œ")
    sp_cg = find_col(spec_single, "ìƒ‰ìƒêµ°")
    if sp_pc and sp_cg and len(spec_single):
        tmp = spec_single[[sp_pc, sp_cg]].dropna()
        tmp[sp_pc] = tmp[sp_pc].apply(normalize_product_code)
        tmp[sp_cg] = tmp[sp_cg].apply(normalize_color_group)
        for pc, g in tmp.groupby(sp_pc)[sp_cg]:
            mapping[str(pc)] = g.value_counts().idxmax()

    s_pc = find_col(single_df, "ì œí’ˆì½”ë“œ")
    s_cg = find_col(single_df, "ìƒ‰ìƒêµ°")
    if s_pc and s_cg and len(single_df):
        tmp = single_df[[s_pc, s_cg]].dropna()
        tmp[s_pc] = tmp[s_pc].apply(normalize_product_code)
        tmp[s_cg] = tmp[s_cg].apply(normalize_color_group)
        for pc, g in tmp.groupby(s_pc)[s_cg]:
            pc = str(pc)
            if pc not in mapping:
                mapping[pc] = g.value_counts().idxmax()

    return mapping

# ==========================================================
# Excel append / download
# ==========================================================
def get_sheet_headers(xlsx_path: str, sheet_name: str) -> list[str]:
    wb = load_workbook(xlsx_path)
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    headers = []
    for cell in ws[1]:
        headers.append(None if cell.value is None else str(cell.value).strip())
    while headers and headers[-1] in (None, "", "nan"):
        headers.pop()
    return headers

def append_row_to_xlsx(xlsx_path: str, sheet_name: str, row_dict: dict):
    wb = load_workbook(xlsx_path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"ì‹œíŠ¸ê°€ ì—†ìŠµë‹ˆë‹¤: {sheet_name}")

    ws = wb[sheet_name]
    headers = get_sheet_headers(xlsx_path, sheet_name)
    if not headers:
        raise ValueError(f"í—¤ë”(1í–‰)ë¥¼ ì°¾ì§€ ëª»í–ˆìŠµë‹ˆë‹¤: {sheet_name}")

    row = []
    for h in headers:
        if h is None:
            row.append(None)
            continue
        v = row_dict.get(h, None)
        if isinstance(v, dt.datetime):
            v = v.date()
        row.append(v)

    ws.append(row)
    wb.save(xlsx_path)

def download_xlsx_button(xlsx_path: str, label: str = "ì—…ë°ì´íŠ¸ëœ ì—‘ì…€ ë‹¤ìš´ë¡œë“œ"):
    try:
        data = Path(xlsx_path).read_bytes()
        st.download_button(
            label,
            data=data,
            file_name=Path(xlsx_path).name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    except Exception as e:
        st.error("ë‹¤ìš´ë¡œë“œ íŒŒì¼ ìƒì„± ì‹¤íŒ¨")
        st.exception(e)

# ==========================================================
# Lot auto generator (ê¸°ì¡´ Lot ìŠ¤íƒ€ì¼ ì¶”ì •)
# ==========================================================
def infer_lot_style(existing_lots: list[str], fallback_prefix: str):
    """
    returns (prefix, date_len, sep, seq_len)
    - prefix: leading letters from last lot (or fallback)
    - date_len: 6(YYMMDD) or 8(YYYYMMDD)
    - sep: "-" or ""
    - seq_len: digits
    """
    last = None
    for x in reversed(existing_lots):
        if x and str(x).strip() and str(x).lower() not in ("nan", "none"):
            last = str(x).strip()
            break

    if not last:
        return (fallback_prefix, 6, "-", 2)

    m = re.match(r"^([A-Za-z]+)(.*)$", last)
    if not m:
        return (fallback_prefix, 6, "-", 2)

    prefix = m.group(1) or fallback_prefix
    rest = (m.group(2) or "").strip()

    sep = "-" if "-" in rest else ""
    if sep:
        parts = rest.split("-", 1)
        date_part = parts[0]
        seq_part = parts[1] if len(parts) > 1 else ""
        date_len = 8 if (date_part.startswith("20") and len(date_part) >= 8) else 6
        seq_len = max(2, len(seq_part)) if seq_part else 2
        return (prefix, date_len, sep, seq_len)

    # no sep
    date_len = 8 if (rest.startswith("20") and len(rest) >= 8) else 6
    seq_len = max(2, len(rest) - date_len)
    return (prefix, date_len, "", seq_len)

def next_lot(existing_lots: list[str], date_value: dt.date, fallback_prefix: str):
    prefix, date_len, sep, seq_len = infer_lot_style(existing_lots, fallback_prefix)
    date_str = date_value.strftime("%Y%m%d") if date_len == 8 else date_value.strftime("%y%m%d")

    pat = re.compile(rf"^{re.escape(prefix)}{re.escape(date_str)}{re.escape(sep)}(\d+)$")
    max_seq = 0
    for x in existing_lots:
        if not x:
            continue
        s = str(x).strip()
        mm = pat.match(s)
        if not mm:
            continue
        try:
            max_seq = max(max_seq, int(mm.group(1)))
        except Exception:
            pass

    seq = max_seq + 1
    seq_str = str(seq).zfill(seq_len)
    return f"{prefix}{date_str}{sep}{seq_str}"

# ==========================================================
# Stock history loader (ì¬ê³  íŒŒì¼: ì‹œíŠ¸ëª… 1.15 ë“±)
# ==========================================================
def _parse_stock_sheet_date(sheet_name: str, today: dt.date):
    s = str(sheet_name).strip()
    m = re.match(r"^(\d{1,2})\.(\d{1,2})$", s)  # ì˜ˆ: 1.15
    if not m:
        return None
    month = int(m.group(1))
    day = int(m.group(2))
    year = today.year
    # ì—°ë§/ì—°ì´ˆ ê²½ê³„ ë³´ì •
    if month > (today.month + 1):
        year -= 1
    try:
        return dt.date(year, month, day)
    except ValueError:
        return None

@st.cache_data(show_spinner=False)
def load_stock_history(stock_xlsx_path: str, stock_sig, product_to_color: dict[str, str]) -> pd.DataFrame:
    p = Path(stock_xlsx_path)
    if not stock_xlsx_path or not p.exists():
        return pd.DataFrame()

    today = dt.date.today()
    xls = pd.ExcelFile(stock_xlsx_path, engine="openpyxl")

    frames = []
    for sh in xls.sheet_names:
        d = _parse_stock_sheet_date(sh, today)
        if d is None:
            continue

        df = pd.read_excel(xls, sheet_name=sh)
        df = df.rename(columns=lambda x: str(x).strip())

        c_div = find_col(df, "êµ¬ë¶„")
        c_item = find_col(df, "í’ˆëª©ëª…")
        c_curr = find_col(df, "ê¸ˆì¼ ì¬ê³ (kg)") or find_col(df, "ê¸ˆì¼ì¬ê³ (kg)") or find_col(df, "ì¬ê³ (kg)")
        c_used = find_col(df, "í•˜ë£¨ ì‚¬ìš©ëŸ‰(kg)") or find_col(df, "ì‚¬ìš©ëŸ‰(kg)") or find_col(df, "ì‚¬ìš©ëŸ‰")

        if not (c_item and c_curr and c_used):
            continue

        out = pd.DataFrame()
        out["division"] = df[c_div].astype(str).str.strip() if c_div else ""
        out["product_code"] = df[c_item].apply(normalize_product_code)

        out["curr_stock_kg"] = pd.to_numeric(
            df[c_curr].astype(str).str.replace(",", "", regex=False), errors="coerce"
        ).fillna(0)

        used_raw = pd.to_numeric(df[c_used].astype(str).str.replace(",", "", regex=False), errors="coerce")
        # ì‚¬ìš©ëŸ‰: ì–‘ìˆ˜ / ì…ê³ (ì¶”ì •): ìŒìˆ˜(ì¬ê³  ì¦ê°€ë¡œ ì…ë ¥ë˜ëŠ” ê²½ìš°) ì²˜ë¦¬
        out["used_kg"] = used_raw.clip(lower=0).fillna(0)
        out["inbound_kg"] = (-used_raw).clip(lower=0).fillna(0)
        out["inbound_event"] = (out["inbound_kg"] > 0).astype(int)

        out = out.dropna(subset=["product_code"])
        out["color_group"] = out["product_code"].map(product_to_color).fillna("OTHER").apply(normalize_color_group)
        out["date"] = pd.to_datetime(d)
        frames.append(out[["date", "division", "product_code", "color_group", "curr_stock_kg", "used_kg", "inbound_kg", "inbound_event"]])

    if not frames:
        return pd.DataFrame()

    hist = pd.concat(frames, ignore_index=True)
    hist = hist.sort_values(["date", "division", "product_code"]).reset_index(drop=True)
    return hist

# ==========================================================
# Google Sheets Reader (public)
# ==========================================================
@st.cache_data(ttl=60, show_spinner=False)
def read_gsheet_csv(sheet_id: str, sheet_name: str) -> pd.DataFrame:
    base = f"https://docs.google.com/spreadsheets/d/{sheet_id}/gviz/tq"
    r = requests.get(base, params={"tqx": "out:csv", "sheet": sheet_name}, timeout=20)
    r.raise_for_status()
    r.encoding = "utf-8"
    return pd.read_csv(StringIO(r.text))

# ==========================================================
# Load Lot excel sheets (Lot ê´€ë¦¬)
# ==========================================================
@st.cache_data(show_spinner=False)
def load_dataframes(xlsx_path: str, lot_sig) -> dict[str, pd.DataFrame]:
    def read(name: str) -> pd.DataFrame:
        return pd.read_excel(xlsx_path, sheet_name=name)

    out = {
        "binder": read(SHEET_BINDER),
        "single": read(SHEET_SINGLE),
        "spec_binder": read(SHEET_SPEC_BINDER),
        "spec_single": read(SHEET_SPEC_SINGLE),
        "base_lab": read(SHEET_BASE_LAB),
    }
    try:
        out["binder_return"] = pd.read_excel(xlsx_path, sheet_name=SHEET_BINDER_RETURN)
    except Exception:
        out["binder_return"] = pd.DataFrame(columns=["ì¼ì", "ë°”ì¸ë”íƒ€ì…", "ë°”ì¸ë”ëª…", "ë°”ì¸ë” Lot", "ë°˜í™˜ëŸ‰(kg)", "ë¹„ê³ "])
    return out

# ==========================================================
# Binder IO file upload (ì—‘ì…€ ì—…ë¡œë“œ ì¦‰ì‹œ í‘œì‹œ)
# ==========================================================
def _guess_hema_sil_sheets(sheet_names: list[str]):
    hema = None
    sil = None
    for s in sheet_names:
        u = str(s).upper()
        if hema is None and ("HEMA" in u or "í—¤ë§ˆ" in str(s)):
            hema = s
        if sil is None and (("SIL" in u) or ("SILIC" in u) or ("ì‹¤ë¦¬" in str(s)) or ("ì‹¤ë¦¬ì½˜" in str(s))):
            sil = s
    return hema, sil

@st.cache_data(show_spinner=False)
def load_binder_io_excel(xlsx_bytes: bytes, filename: str) -> dict[str, pd.DataFrame]:
    tmp = Path(f".binder_io_{re.sub(r'[^A-Za-z0-9_.-]', '_', filename)}")
    tmp.write_bytes(xlsx_bytes)

    xls = pd.ExcelFile(tmp, engine="openpyxl")
    hema_sh, sil_sh = _guess_hema_sil_sheets(xls.sheet_names)

    out = {}
    if hema_sh:
        out["HEMA"] = pd.read_excel(xls, sheet_name=hema_sh)
    if sil_sh:
        out["Silicone"] = pd.read_excel(xls, sheet_name=sil_sh)

    if not out:
        out["ALL"] = pd.read_excel(xls, sheet_name=xls.sheet_names[0])

    # ë‚ ì§œ ì»¬ëŸ¼ ê°ì§€ í›„ ìµœê·¼ìˆœ ì •ë ¬
    for k, df in list(out.items()):
        if df is None or df.empty:
            continue
        dc = detect_date_col(df)
        if dc:
            df2 = df.copy()
            df2["_sort_date"] = pd.to_datetime(df2[dc], errors="coerce")
            df2 = df2.sort_values(by="_sort_date", ascending=False).drop(columns=["_sort_date"])
            out[k] = df2

    return out

# ==========================================================
# Title
# ==========================================================
st.title("ğŸ§ª ì•¡ìƒ ì‰í¬ Lot ì¶”ì  ê´€ë¦¬ ëŒ€ì‹œë³´ë“œ")
st.markdown(
    """
    <div style='background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); 
                padding: 1rem 1.5rem; 
                border-radius: 12px; 
                margin-bottom: 1.5rem;
                color: white;'>
        <h3 style='margin: 0; color: white; font-size: 1.1rem;'>
            ğŸ“Š ì‹¤ì‹œê°„ ì¬ê³  Â· ì ë„ Â· Lot ì¶”ì  í†µí•© ê´€ë¦¬ ì‹œìŠ¤í…œ
        </h3>
        <p style='margin: 0.5rem 0 0 0; font-size: 0.9rem; opacity: 0.9;'>
            âœ… ëŒ€ì‹œë³´ë“œ | âœ… ìš”ì•½ | âœ… ì¬ê³ ê´€ë¦¬ | âœ… ë°”ì¸ë” ì…ì¶œê³  | âœ… ì‹ ê·œ ì…ë ¥ | âœ… ë¹ ë¥¸ê²€ìƒ‰
        </p>
    </div>
    """,
    unsafe_allow_html=True
)

# ==========================================================
# Sidebar - files
# ==========================================================
with st.sidebar:
    st.header("ë°ì´í„° íŒŒì¼ (Lot ê´€ë¦¬)")
    xlsx_path = st.text_input("ì—‘ì…€ íŒŒì¼ ê²½ë¡œ", value=DEFAULT_XLSX)
    uploaded = st.file_uploader("ë˜ëŠ” ì—‘ì…€ ì—…ë¡œë“œ(.xlsx)", type=["xlsx"], key="lot_upload")

    st.divider()
    st.header("ì¬ê³  íŒŒì¼")
    stock_xlsx_path = st.text_input("ì¬ê³  ì—‘ì…€ íŒŒì¼ ê²½ë¡œ", value=DEFAULT_STOCK_XLSX, key="stock_path")
    uploaded_stock = st.file_uploader("ë˜ëŠ” ì¬ê³  ì—‘ì…€ ì—…ë¡œë“œ(.xlsx)", type=["xlsx"], key="stock_upload")

# ì—…ë¡œë“œ íŒŒì¼ì„ ì„ì‹œ íŒŒì¼ë¡œ ì‚¬ìš©(ì „ì²´ êµì²´ìš©)
if uploaded is not None:
    sig = f"{uploaded.name}:{uploaded.size}"
    if st.session_state.get("_uploaded_sig") != sig:
        tmp_path = Path(".streamlit_tmp.xlsx")
        tmp_path.write_bytes(uploaded.getvalue())
        st.session_state["_uploaded_sig"] = sig
        st.session_state["_tmp_xlsx_path"] = str(tmp_path)
    xlsx_path = st.session_state.get("_tmp_xlsx_path", xlsx_path)
    st.sidebar.info("ì—…ë¡œë“œ íŒŒì¼(Lot ê´€ë¦¬)ë¡œ ì‹¤í–‰ ì¤‘ì…ë‹ˆë‹¤. (ì›ë³¸ ìë™ ì €ì¥ì´ ì•„ë‹ˆë¼, ì—…ë°ì´íŠ¸ í›„ 'ë‹¤ìš´ë¡œë“œ'ë¡œ ë°›ëŠ” ë°©ì‹ì´ ì•ˆì „í•©ë‹ˆë‹¤.)")

if uploaded_stock is not None:
    sig = f"{uploaded_stock.name}:{uploaded_stock.size}"
    if st.session_state.get("_uploaded_sig_stock") != sig:
        tmp_path = Path(".streamlit_tmp_stock.xlsx")
        tmp_path.write_bytes(uploaded_stock.getvalue())
        st.session_state["_uploaded_sig_stock"] = sig
        st.session_state["_tmp_stock_path"] = str(tmp_path)
    stock_xlsx_path = st.session_state.get("_tmp_stock_path", stock_xlsx_path)
    st.sidebar.info("ì—…ë¡œë“œ íŒŒì¼(ì¬ê³ )ë¡œ ì‹¤í–‰ ì¤‘ì…ë‹ˆë‹¤.")

# ==========================================================
# Load Lot excel (ì—†ìœ¼ë©´ ë¹ˆ ë°ì´í„°ë¼ë„ í™”ë©´ í‘œì‹œ)
# ==========================================================
lot_sig = file_sig(xlsx_path)
if not Path(xlsx_path).exists():
    st.error(f"ì—‘ì…€ íŒŒì¼ì„ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤: {xlsx_path}")
    st.info("ì¢Œì¸¡ ì‚¬ì´ë“œë°”ì—ì„œ Lotê´€ë¦¬ ì—‘ì…€ì„ ì—…ë¡œë“œí•˜ê±°ë‚˜, ê²½ë¡œë¥¼ ì˜¬ë°”ë¥´ê²Œ ìˆ˜ì •í•´ì£¼ì„¸ìš”. (í˜„ì¬ëŠ” 'ë¹ˆ ë°ì´í„°'ë¡œ í™”ë©´ë§Œ í‘œì‹œí•©ë‹ˆë‹¤.)")

    binder_df = pd.DataFrame(columns=["ì œì¡°/ì…ê³ ì¼", "Lot(ìë™)", "íŒì •"])
    single_df = pd.DataFrame(columns=["ì…ê³ ì¼","ì ë„ì¸¡ì •ê°’(cP)","ì ë„íŒì •","ë‹¨ì¼ìƒ‰ì‰í¬ Lot","ì‚¬ìš©ëœ ë°”ì¸ë” Lot","ìƒ‰ìƒêµ°","ì œí’ˆì½”ë“œ"])
    spec_binder = pd.DataFrame()
    spec_single = pd.DataFrame(columns=["ì œí’ˆì½”ë“œ", "ìƒ‰ìƒêµ°"])
    base_lab = pd.DataFrame()
    binder_return_df = pd.DataFrame(columns=["ì¼ì", "ë°”ì¸ë”íƒ€ì…", "ë°”ì¸ë”ëª…", "ë°”ì¸ë” Lot", "ë°˜í™˜ëŸ‰(kg)", "ë¹„ê³ "])
else:
    ensure_sheet_exists(
        xlsx_path,
        SHEET_BINDER_RETURN,
        headers=["ì¼ì", "ë°”ì¸ë”íƒ€ì…", "ë°”ì¸ë”ëª…", "ë°”ì¸ë” Lot", "ë°˜í™˜ëŸ‰(kg)", "ë¹„ê³ "],
    )
    data = load_dataframes(xlsx_path, lot_sig)
    binder_df = data["binder"].copy()
    single_df = data["single"].copy()
    spec_binder = data["spec_binder"].copy()
    spec_single = data["spec_single"].copy()
    base_lab = data["base_lab"].copy()
    binder_return_df = data["binder_return"].copy()

# normalize dates (Lot)
c_b_date = find_col(binder_df, "ì œì¡°/ì…ê³ ì¼")
c_s_date = find_col(single_df, "ì…ê³ ì¼")
if c_b_date and c_b_date in binder_df.columns:
    binder_df[c_b_date] = binder_df[c_b_date].apply(normalize_date)
if c_s_date and c_s_date in single_df.columns:
    single_df[c_s_date] = single_df[c_s_date].apply(normalize_date)

# common cols (Single)
c_s_visc = find_col(single_df, "ì ë„ì¸¡ì •ê°’(cP)")
c_s_judge = find_col(single_df, "ì ë„íŒì •")
c_s_lot = find_col(single_df, "ë‹¨ì¼ìƒ‰ì‰í¬ Lot")
c_s_blot = find_col(single_df, "ì‚¬ìš©ëœ ë°”ì¸ë” Lot")
c_s_cg = find_col(single_df, "ìƒ‰ìƒêµ°")
c_s_pc = find_col(single_df, "ì œí’ˆì½”ë“œ")

# ==========================================================
# Tabs
# ==========================================================
tab_dash, tab_summary, tab_stock, tab_binder, tab_input, tab_search = st.tabs(
    ["ğŸ“Š ëŒ€ì‹œë³´ë“œ", "ğŸ“Œ ìš”ì•½", "ğŸ“¦ ì•¡ìƒì‰í¬ ì¬ê³ ê´€ë¦¬", "ğŸ“¦ ë°”ì¸ë” ì…ì¶œê³ ", "ğŸ“ ì‹ ê·œ ì…ë ¥", "ğŸ” ë¹ ë¥¸ê²€ìƒ‰"]
)

# ==========================================================
# Render: Summary (ìƒì‚¬ìš© 1ì¥ ìš”ì•½)
# ==========================================================
def render_summary():
    st.markdown('<div class="section-title">ğŸ“Œ ê²½ì˜ì§„ ìš”ì•½ ë¦¬í¬íŠ¸</div>', unsafe_allow_html=True)
    st.markdown('<div class="section-sub">í•µì‹¬ KPIì™€ ì£¼ìš” ì§€í‘œë¥¼ í•œëˆˆì— í™•ì¸í•  ìˆ˜ ìˆìŠµë‹ˆë‹¤</div>', unsafe_allow_html=True)

    stock_ok = bool(stock_xlsx_path and Path(stock_xlsx_path).exists())
    product_to_color = build_product_to_color_map(spec_single, single_df)

    # ---------- ì¬ê³  KPI ----------
    inv_color = pd.DataFrame()
    use_color = pd.DataFrame()
    cov_alert = pd.DataFrame()
    stock_kpis = {}

    if stock_ok:
        stock_sig = file_sig(stock_xlsx_path)
        hist = load_stock_history(stock_xlsx_path, stock_sig, product_to_color)
        if not hist.empty:
            max_d = hist["date"].max().date()
            start = max(hist["date"].min().date(), max_d - dt.timedelta(days=29))
            end = max_d
            day_span = max(1, (end - start).days + 1)

            hist_f = hist[(hist["date"].dt.date >= start) & (hist["date"].dt.date <= end)].copy()
            latest_df = hist[hist["date"].dt.date == end].copy()

            stock_kpis["ì¬ê³  ìµœì‹ ì¼"] = end.isoformat()
            stock_kpis["í˜„ì¬ ì´ ì¬ê³ (kg)"] = float(latest_df["curr_stock_kg"].sum())
            stock_kpis["ìµœê·¼ 30ì¼ ì‚¬ìš©ëŸ‰(kg)"] = float(hist_f["used_kg"].sum())
            stock_kpis["ìµœê·¼ 30ì¼ ì…ê³ (ê±´)"] = int(hist_f["inbound_event"].sum())
            stock_kpis["í‰ê·  ì‚¬ìš©ëŸ‰(kg/ì¼)"] = float(stock_kpis["ìµœê·¼ 30ì¼ ì‚¬ìš©ëŸ‰(kg)"] / day_span)

            inv_color = (
                latest_df.groupby("color_group", as_index=False)["curr_stock_kg"]
                .sum().rename(columns={"curr_stock_kg": "kg"})
                .sort_values("kg", ascending=False)
            )
            use_color = (
                hist_f.groupby("color_group", as_index=False)["used_kg"]
                .sum().rename(columns={"used_kg": "kg"})
                .sort_values("kg", ascending=False)
            )

            # ì»¤ë²„ë¦¬ì§€ (stock / avg daily use)
            use_by_product = hist_f.groupby("product_code", as_index=False)["used_kg"].sum()
            use_by_product["avg_daily_use"] = use_by_product["used_kg"] / day_span

            stock_by_product = latest_df.groupby("product_code", as_index=False)["curr_stock_kg"].sum().rename(
                columns={"curr_stock_kg": "stock_kg"}
            )

            cov = stock_by_product.merge(use_by_product[["product_code", "avg_daily_use"]], on="product_code", how="left")
            cov["avg_daily_use"] = cov["avg_daily_use"].fillna(0.0)
            cov["cover_days"] = cov.apply(
                lambda r: (r["stock_kg"] / r["avg_daily_use"]) if r["avg_daily_use"] > 0 else None, axis=1
            )
            cov_alert = cov[cov["cover_days"].notna()].sort_values("cover_days").head(10)
        else:
            stock_ok = False

    # ---------- ì ë„ KPI ----------
    visc_ok = bool(
        c_s_date and c_s_visc and c_s_pc
        and (c_s_date in single_df.columns)
        and (c_s_visc in single_df.columns)
        and (c_s_pc in single_df.columns)
    )

    visc_kpis = {}
    daily_visc = pd.DataFrame()
    top_ng = pd.DataFrame()

    if visc_ok:
        df = single_df.copy()
        df[c_s_date] = pd.to_datetime(df[c_s_date], errors="coerce")
        df["_ì ë„"] = pd.to_numeric(df[c_s_visc].astype(str).str.replace(",", "", regex=False), errors="coerce")
        df[c_s_pc] = df[c_s_pc].apply(normalize_product_code)

        df = df.dropna(subset=[c_s_date, "_ì ë„", c_s_pc])
        if len(df):
            max_d = df[c_s_date].max().date()
            start = max(df[c_s_date].min().date(), max_d - dt.timedelta(days=29))
            df30 = df[(df[c_s_date].dt.date >= start) & (df[c_s_date].dt.date <= max_d)].copy()

            total = len(df30)
            ng = int((df30[c_s_judge] == "ë¶€ì í•©").sum()) if c_s_judge and (c_s_judge in df30.columns) else 0
            ng_rate = (ng / total * 100) if total else 0.0

            visc_kpis = {
                "ì ë„ ìµœì‹ ì¼": max_d.isoformat(),
                "ìµœê·¼ 30ì¼ ì¸¡ì •(ê±´)": total,
                "ë¶€ì í•©(ê±´)": ng,
                "ë¶€ì í•©ë¥ (%)": ng_rate,
            }

            # âœ… ìˆ˜ì •ëœ ë¶€ë¶„: groupby ì˜¤ë¥˜ í•´ê²°
            daily_visc = (
                df30.groupby(df30[c_s_date].dt.date)
                .agg(mean_visc=("_ì ë„", "mean"), cnt=("_ì ë„", "size"))
                .reset_index()
                .rename(columns={c_s_date: "date"})
            )
            daily_visc["date"] = pd.to_datetime(daily_visc["date"])

            if c_s_judge and (c_s_judge in df30.columns):
                top_ng = (
                    df30[df30[c_s_judge] == "ë¶€ì í•©"]
                    .groupby(c_s_pc).size().reset_index(name="ng_cnt")
                    .sort_values("ng_cnt", ascending=False).head(8)
                )
        else:
            visc_ok = False

    # ---------- KPI Row ----------
    st.markdown("### ğŸ“Š í•µì‹¬ ì„±ê³¼ ì§€í‘œ (KPI)")
    
    a, b = st.columns(2)
    with a:
        st.markdown("#### ğŸ“¦ ì¬ê³  í˜„í™© (ìµœê·¼ 30ì¼)")
        if not stock_ok:
            st.info("ğŸ’¡ ì¬ê³  íŒŒì¼ì´ ì—†ê±°ë‚˜ ì½ì§€ ëª»í–ˆìŠµë‹ˆë‹¤. ì¢Œì¸¡ ì‚¬ì´ë“œë°”ì—ì„œ ì¬ê³  íŒŒì¼ì„ ì„¤ì •í•´ì£¼ì„¸ìš”.")
        else:
            k1, k2 = st.columns(2)
            k1.metric("ğŸ“… ìµœì‹  ì—…ë°ì´íŠ¸", stock_kpis["ì¬ê³  ìµœì‹ ì¼"])
            k2.metric("ğŸ“Š ì´ ì¬ê³ ëŸ‰", f'{stock_kpis["í˜„ì¬ ì´ ì¬ê³ (kg)"]:,.1f} kg')
            
            k3, k4, k5 = st.columns(3)
            k3.metric("ğŸ“‰ 30ì¼ ì‚¬ìš©ëŸ‰", f'{stock_kpis["ìµœê·¼ 30ì¼ ì‚¬ìš©ëŸ‰(kg)"]:,.1f} kg')
            k4.metric("ğŸ“¥ ì…ê³  ê±´ìˆ˜", f'{stock_kpis["ìµœê·¼ 30ì¼ ì…ê³ (ê±´)"]:,}')
            k5.metric("âš¡ ì¼í‰ê·  ì‚¬ìš©", f'{stock_kpis["í‰ê·  ì‚¬ìš©ëŸ‰(kg/ì¼)"]:,.1f} kg/ì¼')

    with b:
        st.markdown("#### ğŸ§ª í’ˆì§ˆ í˜„í™© (ìµœê·¼ 30ì¼)")
        if not visc_ok:
            st.info("ğŸ’¡ ë‹¨ì¼ìƒ‰ ì‹œíŠ¸ì— ì…ê³ ì¼/ì ë„ì¸¡ì •ê°’/ì œí’ˆì½”ë“œ ì»¬ëŸ¼ì´ í•„ìš”í•©ë‹ˆë‹¤.")
        else:
            k1, k2 = st.columns(2)
            k1.metric("ğŸ“… ìµœì‹  ì¸¡ì •ì¼", visc_kpis["ì ë„ ìµœì‹ ì¼"])
            k2.metric("ğŸ”¬ ì´ ì¸¡ì • ê±´ìˆ˜", f'{visc_kpis["ìµœê·¼ 30ì¼ ì¸¡ì •(ê±´)"]:,}')
            
            k3, k4 = st.columns(2)
            k3.metric("âŒ ë¶€ì í•©", f'{visc_kpis["ë¶€ì í•©(ê±´)"]:,} ê±´')
            
            ng_rate = visc_kpis["ë¶€ì í•©ë¥ (%)"]
            if ng_rate > 10:
                k4.markdown(f'<div class="status-badge-error">âš ï¸ ë¶€ì í•©ë¥  {ng_rate:.1f}%</div>', unsafe_allow_html=True)
            elif ng_rate > 5:
                k4.markdown(f'<div class="status-badge-warning">âš ï¸ ë¶€ì í•©ë¥  {ng_rate:.1f}%</div>', unsafe_allow_html=True)
            else:
                k4.markdown(f'<div class="status-badge-success">âœ… ë¶€ì í•©ë¥  {ng_rate:.1f}%</div>', unsafe_allow_html=True)

    st.divider()
    st.markdown("### ğŸ“ˆ í•µì‹¬ ì‹œê°í™” ì°¨íŠ¸")

    c1, c2 = st.columns(2)
    with c1:
        st.markdown('<div class="chart-container">', unsafe_allow_html=True)
        st.markdown("**ğŸ“¦ í˜„ì¬ ì¬ê³  í˜„í™© (ìƒ‰ìƒë³„)**")
        if stock_ok and not inv_color.empty:
            ch = alt.Chart(inv_color).mark_bar().encode(
                y=alt.Y("color_group:N", sort="-x", title="ìƒ‰ìƒ ê³„ì—´"),
                x=alt.X("kg:Q", title="ì¬ê³ ëŸ‰ (kg)"),
                color=alt.Color("color_group:N", scale=_color_scale_color_group(), legend=None),
                tooltip=[alt.Tooltip("color_group:N", title="ìƒ‰ìƒê³„ì—´"), alt.Tooltip("kg:Q", title="ì¬ê³ (kg)", format=",.1f")],
            ).properties(height=280)
            st.altair_chart(ch, use_container_width=True)
        else:
            st.info("ğŸ“Š ì¬ê³  ë°ì´í„°ê°€ ì—†ìŠµë‹ˆë‹¤.")
        st.markdown('</div>', unsafe_allow_html=True)

    with c2:
        st.markdown('<div class="chart-container">', unsafe_allow_html=True)
        st.markdown("**ğŸ§ª ì ë„ ì¶”ì´ (ì¼ë³„ í‰ê· )**")
        if visc_ok and not daily_visc.empty:
            ch = alt.Chart(daily_visc).mark_line(point=True, strokeWidth=3).encode(
                x=alt.X("date:T", title="ë‚ ì§œ"),
                y=alt.Y("mean_visc:Q", title="í‰ê·  ì ë„ (cP)"),
                tooltip=[
                    alt.Tooltip("date:T", title="ë‚ ì§œ"),
                    alt.Tooltip("mean_visc:Q", title="í‰ê· ì ë„", format=",.0f"),
                    alt.Tooltip("cnt:Q", title="ì¸¡ì •(ê±´)", format=",.0f"),
                ],
            ).properties(height=280)
            st.altair_chart(ch, use_container_width=True)
        else:
            st.info("ğŸ“Š ì ë„ ë°ì´í„°ê°€ ì—†ìŠµë‹ˆë‹¤.")
        st.markdown('</div>', unsafe_allow_html=True)

    c3, c4 = st.columns(2)
    with c3:
        st.markdown('<div class="chart-container">', unsafe_allow_html=True)
        st.markdown("**ğŸ“‰ 30ì¼ ì‚¬ìš©ëŸ‰ (ìƒ‰ìƒë³„)**")
        if stock_ok and not use_color.empty:
            ch = alt.Chart(use_color).mark_bar().encode(
                y=alt.Y("color_group:N", sort="-x", title="ìƒ‰ìƒ ê³„ì—´"),
                x=alt.X("kg:Q", title="ì‚¬ìš©ëŸ‰ (kg)"),
                color=alt.Color("color_group:N", scale=_color_scale_color_group(), legend=None),
                tooltip=[alt.Tooltip("color_group:N", title="ìƒ‰ìƒê³„ì—´"), alt.Tooltip("kg:Q", title="ì‚¬ìš©ëŸ‰(kg)", format=",.1f")],
            ).properties(height=280)
            st.altair_chart(ch, use_container_width=True)
        else:
            st.info("ğŸ“Š ì‚¬ìš©ëŸ‰ ë°ì´í„°ê°€ ì—†ìŠµë‹ˆë‹¤.")
        st.markdown('</div>', unsafe_allow_html=True)

    with c4:
        st.markdown('<div class="chart-container">', unsafe_allow_html=True)
        st.markdown("**âš ï¸ ë¶€ì í•© ë‹¤ë°œ ì œí’ˆ (Top 8)**")
        if visc_ok and not top_ng.empty:
            ch = alt.Chart(top_ng).mark_bar(color='#ef4444').encode(
                y=alt.Y(f"{c_s_pc}:N", sort="-x", title="ì œí’ˆ ì½”ë“œ"),
                x=alt.X("ng_cnt:Q", title="ë¶€ì í•© ê±´ìˆ˜"),
                tooltip=[alt.Tooltip(f"{c_s_pc}:N", title="ì œí’ˆì½”ë“œ"), alt.Tooltip("ng_cnt:Q", title="ë¶€ì í•©(ê±´)", format=",.0f")],
            ).properties(height=280)
            st.altair_chart(ch, use_container_width=True)
        else:
            st.success("âœ… ë¶€ì í•© ë°ì´í„°ê°€ ì—†ìŠµë‹ˆë‹¤.")
        st.markdown('</div>', unsafe_allow_html=True)

    with st.expander("ğŸ” ìƒì„¸ ë¶„ì„: ì¬ê³  ë¶€ì¡± ê²½ë³´ (ì»¤ë²„ë¦¬ì§€ Top 10)"):
        if stock_ok and not cov_alert.empty:
            st.warning("âš ï¸ ë‹¤ìŒ ì œí’ˆë“¤ì€ ì¬ê³  ì†Œì§„ ìœ„í—˜ì´ ìˆìŠµë‹ˆë‹¤. ë°œì£¼ë¥¼ ê²€í† í•´ì£¼ì„¸ìš”.")
            show = cov_alert.copy()
            show["stock_kg"] = show["stock_kg"].round(1)
            show["avg_daily_use"] = show["avg_daily_use"].round(2)
            show["cover_days"] = show["cover_days"].round(1)
            st.dataframe(show, use_container_width=True, height=320)
        else:
            st.success("âœ… í˜„ì¬ ì¬ê³  ë¶€ì¡± ìœ„í—˜ ì œí’ˆì´ ì—†ìŠµë‹ˆë‹¤.")

# ==========================================================
# Render: Stock tab (ì¬ê³ /ì…ê³ /ì‚¬ìš©ëŸ‰ì„ í•œ íƒ­ì—ì„œ ë³´ê¸° ì¢‹ê²Œ)
# ==========================================================
def render_stock_tab():
    st.markdown('<div class="section-title">ğŸ“¦ ì•¡ìƒì‰í¬ ì¬ê³ ê´€ë¦¬</div>', unsafe_allow_html=True)
    st.markdown('<div class="section-sub">ì¬ê³  í˜„í™© Â· ì…ê³  ì¶”ì • Â· ì‚¬ìš©ëŸ‰ ì¶”ì´ë¥¼ ìƒ‰ìƒë³„ë¡œ ë¶„ì„í•©ë‹ˆë‹¤</div>', unsafe_allow_html=True)

    if not stock_xlsx_path or not Path(stock_xlsx_path).exists():
        st.error("âŒ ì¬ê³  íŒŒì¼ ê²½ë¡œê°€ ì˜¬ë°”ë¥´ì§€ ì•ŠìŠµë‹ˆë‹¤. ì¢Œì¸¡ ì‚¬ì´ë“œë°”ì—ì„œ ì¬ê³  íŒŒì¼ì„ ì„¤ì •í•´ì£¼ì„¸ìš”.")
        return

    product_to_color = build_product_to_color_map(spec_single, single_df)
    stock_sig = file_sig(stock_xlsx_path)
    hist = load_stock_history(stock_xlsx_path, stock_sig, product_to_color)
    if hist.empty:
        st.error("âŒ ì¬ê³  ì—‘ì…€ì„ ì½ì§€ ëª»í–ˆìŠµë‹ˆë‹¤. (ì‹œíŠ¸ëª…: 1.15 í˜•ì‹ / ì»¬ëŸ¼: í’ˆëª©ëª…, ê¸ˆì¼ ì¬ê³ (kg), í•˜ë£¨ ì‚¬ìš©ëŸ‰(kg) í™•ì¸)")
        return

    min_d = hist["date"].min().date()
    max_d = hist["date"].max().date()

    # í•„í„° ì„¹ì…˜
    st.markdown("### ğŸ” ì¡°íšŒ ê¸°ê°„ ë° í•„í„°")
    left, mid, right = st.columns([2.2, 2.8, 5.0])
    with left:
        quick = st.selectbox("ğŸ“… ê¸°ê°„ ì„ íƒ", ["ìµœê·¼ 7ì¼", "ìµœê·¼ 30ì¼", "ìµœê·¼ 90ì¼", "ì „ì²´", "ì§ì ‘ ì„ íƒ"], index=1)
    with mid:
        if quick == "ì§ì ‘ ì„ íƒ":
            start = st.date_input("ì‹œì‘ì¼", value=max(min_d, max_d - dt.timedelta(days=30)), min_value=min_d, max_value=max_d)
            end = st.date_input("ì¢…ë£Œì¼", value=max_d, min_value=min_d, max_value=max_d)
        else:
            if quick == "ìµœê·¼ 7ì¼":
                start = max(min_d, max_d - dt.timedelta(days=6))
            elif quick == "ìµœê·¼ 30ì¼":
                start = max(min_d, max_d - dt.timedelta(days=29))
            elif quick == "ìµœê·¼ 90ì¼":
                start = max(min_d, max_d - dt.timedelta(days=89))
            else:
                start = min_d
            end = max_d
            st.write(f"**ğŸ“… {start} ~ {end}**")
    with right:
        divisions = sorted([x for x in hist["division"].dropna().unique().tolist() if str(x).strip() and str(x).lower() not in ("nan", "none")])
        sel_div = st.multiselect("ğŸ­ ì œí’ˆêµ° (PL/NPL/NSL ë“±)", divisions, default=divisions)

    if start > end:
        start, end = end, start

    filt = (hist["date"].dt.date >= start) & (hist["date"].dt.date <= end)
    if sel_div:
        filt = filt & (hist["division"].isin(sel_div))
    hist_f = hist[filt].copy()

    latest_date = hist["date"].max()
    latest_df = hist[hist["date"] == latest_date].copy()
    if sel_div:
        latest_df = latest_df[latest_df["division"].isin(sel_div)].copy()

    total_stock = float(latest_df["curr_stock_kg"].sum())
    total_used = float(hist_f["used_kg"].sum())
    inbound_events = int(hist_f["inbound_event"].sum())
    inbound_kg = float(hist_f["inbound_kg"].sum())
    day_span = max(1, (end - start).days + 1)
    avg_daily_use = total_used / day_span if day_span else 0.0

    # KPI ì¹´ë“œ
    st.markdown("### ğŸ“Š ì£¼ìš” ì§€í‘œ")
    k1, k2, k3, k4, k5, k6 = st.columns(6)
    k1.metric("ğŸ“… ìµœì‹  ì—…ë°ì´íŠ¸", latest_date.date().isoformat())
    k2.metric("ğŸ“¦ í˜„ì¬ ì´ ì¬ê³ ", f"{total_stock:,.1f} kg")
    k3.metric("ğŸ“‰ ê¸°ê°„ ì‚¬ìš©ëŸ‰", f"{total_used:,.1f} kg")
    k4.metric("ğŸ“¥ ì…ê³  ê±´ìˆ˜", f"{inbound_events:,}")
    k5.metric("ğŸ“¦ ì…ê³ ëŸ‰", f"{inbound_kg:,.1f} kg")
    k6.metric("âš¡ ì¼í‰ê·  ì‚¬ìš©", f"{avg_daily_use:,.1f} kg")

    st.markdown('<div class="kpi-note">ğŸ’¡ ì…ê³ ëŠ” "í•˜ë£¨ ì‚¬ìš©ëŸ‰"ì´ ìŒìˆ˜ë¡œ ê¸°ì…ëœ ê²½ìš°(ì¬ê³  ì¦ê°€)ë¥¼ ì…ê³ ë¡œ ì¶”ì •í•©ë‹ˆë‹¤.</div>', unsafe_allow_html=True)
    st.divider()

    # ---------- 1) ìƒ‰ìƒê³„ì—´ ë°”ì°¨íŠ¸ ----------
    inv = latest_df.groupby("color_group", as_index=False)["curr_stock_kg"].sum().rename(columns={"curr_stock_kg": "kg"}).sort_values("kg", ascending=False)
    use = hist_f.groupby("color_group", as_index=False)["used_kg"].sum().rename(columns={"used_kg": "kg"}).sort_values("kg", ascending=False)
    inbound = hist_f.groupby("color_group", as_index=False)["inbound_kg"].sum().rename(columns={"inbound_kg": "kg"}).sort_values("kg", ascending=False)

    def bar_chart(df: pd.DataFrame, value_title: str):
        if df.empty:
            return None
        
        # ë§‰ëŒ€ ì°¨íŠ¸
        bars = alt.Chart(df).mark_bar().encode(
            y=alt.Y("color_group:N", sort="-x", title=""),
            x=alt.X("kg:Q", title=value_title),
            color=alt.Color("color_group:N", scale=_color_scale_color_group(), legend=None),
            tooltip=[alt.Tooltip("color_group:N", title="ìƒ‰ìƒê³„ì—´"), alt.Tooltip("kg:Q", title=value_title, format=",.1f")],
        )
        
        # í…ìŠ¤íŠ¸ ë ˆì´ë¸” (ë§‰ëŒ€ ëì— ê°’ í‘œì‹œ)
        text = alt.Chart(df).mark_text(
            align='left',
            baseline='middle',
            dx=3,  # ë§‰ëŒ€ ëì—ì„œ 3px ì˜¤ë¥¸ìª½
            fontSize=11,
            fontWeight='bold'
        ).encode(
            y=alt.Y("color_group:N", sort="-x", title=""),
            x=alt.X("kg:Q"),
            text=alt.Text("kg:Q", format=",.1f"),
            color=alt.value('#333333')
        )
        
        return (bars + text).properties(height=260)

    st.markdown("### ğŸ“Š ìƒ‰ìƒë³„ ì¬ê³  Â· ì‚¬ìš©ëŸ‰ Â· ì…ê³  í˜„í™©")
    c1, c2, c3 = st.columns(3)
    with c1:
        st.markdown('<div class="chart-container">', unsafe_allow_html=True)
        st.markdown("**1) í˜„ì¬ ì¬ê³  (ìµœì‹ ì¼)**")
        ch = bar_chart(inv, "ì¬ê³ (kg)")
        if ch is not None:
            st.altair_chart(ch, use_container_width=True)
        else:
            st.info("ğŸ“Š í‘œì‹œí•  ì¬ê³  ë°ì´í„°ê°€ ì—†ìŠµë‹ˆë‹¤.")
        st.markdown('</div>', unsafe_allow_html=True)
        
    with c2:
        st.markdown('<div class="chart-container">', unsafe_allow_html=True)
        st.markdown("**2) ê¸°ê°„ ì‚¬ìš©ëŸ‰**")
        ch = bar_chart(use, "ì‚¬ìš©ëŸ‰(kg)")
        if ch is not None:
            st.altair_chart(ch, use_container_width=True)
        else:
            st.info("ğŸ“Š í‘œì‹œí•  ì‚¬ìš©ëŸ‰ ë°ì´í„°ê°€ ì—†ìŠµë‹ˆë‹¤.")
        st.markdown('</div>', unsafe_allow_html=True)
        
    with c3:
        st.markdown('<div class="chart-container">', unsafe_allow_html=True)
        st.markdown("**3) ê¸°ê°„ ì…ê³  (ì¶”ì •)**")
        ch = bar_chart(inbound, "ì…ê³ (kg)")
        if ch is not None:
            st.altair_chart(ch, use_container_width=True)
        else:
            st.info("ğŸ“Š í‘œì‹œí•  ì…ê³  ë°ì´í„°ê°€ ì—†ìŠµë‹ˆë‹¤.")
        st.markdown('</div>', unsafe_allow_html=True)

    st.divider()

    # ---------- 2) ì¼ë³„ ì‚¬ìš©ëŸ‰ ì¶”ì´ ----------
    st.markdown("### ğŸ“ˆ ì¼ë³„ ì‚¬ìš©ëŸ‰ ì¶”ì´ ë¶„ì„")
    present = [k for k in COLOR_KEYS if k in hist_f["color_group"].unique().tolist()]
    default_keys = [k for k in present if k != "OTHER"][:5] or present
    sel_keys = st.multiselect("ğŸ¨ í‘œì‹œí•  ìƒ‰ìƒê³„ì—´", COLOR_KEYS, default=default_keys)

    daily = (
        hist_f[hist_f["color_group"].isin(sel_keys)]
        .groupby(["date", "color_group"], as_index=False)["used_kg"].sum()
    )
    total = hist_f.groupby("date", as_index=False)["used_kg"].sum().rename(columns={"used_kg": "TOTAL"})

    st.markdown('<div class="chart-container">', unsafe_allow_html=True)
    line = alt.Chart(daily).mark_line(point=True, strokeWidth=2.5).encode(
        x=alt.X("date:T", title="ë‚ ì§œ"),
        y=alt.Y("used_kg:Q", title="ì‚¬ìš©ëŸ‰(kg)"),
        color=alt.Color("color_group:N", scale=_color_scale_color_group(), legend=alt.Legend(title="ìƒ‰ìƒê³„ì—´")),
        tooltip=[
            alt.Tooltip("date:T", title="ë‚ ì§œ"),
            alt.Tooltip("color_group:N", title="ìƒ‰ìƒê³„ì—´"),
            alt.Tooltip("used_kg:Q", title="ì‚¬ìš©ëŸ‰(kg)", format=",.1f"),
        ],
    )
    total_line = alt.Chart(total).mark_line(point=True, strokeDash=[6, 3], strokeWidth=3, color='#374151').encode(
        x="date:T",
        y=alt.Y("TOTAL:Q", title="ì‚¬ìš©ëŸ‰(kg)"),
        tooltip=[alt.Tooltip("date:T", title="ë‚ ì§œ"), alt.Tooltip("TOTAL:Q", title="TOTAL(kg)", format=",.1f")],
    )
    st.altair_chart((line + total_line).interactive(), use_container_width=True)
    st.markdown('</div>', unsafe_allow_html=True)

    st.divider()

    # ---------- 3) ì»¤ë²„ë¦¬ì§€(ë°œì£¼ íŒë‹¨ìš©) ----------
    st.markdown("### âš ï¸ ì¬ê³  ì»¤ë²„ë¦¬ì§€ ë¶„ì„ (ë°œì£¼ íŒë‹¨)")
    st.caption("ğŸ’¡ ì»¤ë²„ë¦¬ì§€ = (ìµœì‹  ì¬ê³  kg) / (ì„ íƒê¸°ê°„ í‰ê·  ì¼ì‚¬ìš©ëŸ‰). í‰ê·  ì‚¬ìš©ëŸ‰ì´ 0ì´ë©´ ì»¤ë²„ë¦¬ì§€ ê³„ì‚° ì œì™¸.")

    use_by_product = hist_f.groupby("product_code", as_index=False)["used_kg"].sum()
    use_by_product["avg_daily_use"] = use_by_product["used_kg"] / day_span
    stock_by_product = latest_df.groupby("product_code", as_index=False)["curr_stock_kg"].sum().rename(columns={"curr_stock_kg": "stock_kg"})

    cov = stock_by_product.merge(use_by_product[["product_code", "avg_daily_use"]], on="product_code", how="left")
    cov["avg_daily_use"] = cov["avg_daily_use"].fillna(0.0)
    cov["cover_days"] = cov.apply(lambda r: (r["stock_kg"] / r["avg_daily_use"]) if r["avg_daily_use"] > 0 else None, axis=1)
    cov["color_group"] = cov["product_code"].map(product_to_color).fillna("OTHER").apply(normalize_color_group)

    warn_days = st.slider("âš ï¸ ê²½ë³´ ê¸°ì¤€ (ì¼)", min_value=1, max_value=60, value=14, step=1)
    show_cov = cov[cov["cover_days"].notna()].copy()
    show_cov["cover_days"] = show_cov["cover_days"].round(1)
    show_cov["stock_kg"] = show_cov["stock_kg"].round(1)
    show_cov["avg_daily_use"] = show_cov["avg_daily_use"].round(2)

    alert = show_cov[show_cov["cover_days"] <= warn_days].sort_values("cover_days").head(30)
    if alert.empty:
        st.success(f"âœ… ì»¤ë²„ë¦¬ì§€ {warn_days}ì¼ ì´í•˜ í’ˆëª©ì´ ì—†ìŠµë‹ˆë‹¤. ì¬ê³  ìƒíƒœê°€ ì–‘í˜¸í•©ë‹ˆë‹¤.")
    else:
        st.warning(f"âš ï¸ ì»¤ë²„ë¦¬ì§€ {warn_days}ì¼ ì´í•˜ í’ˆëª©ì´ **{len(alert)}ê°œ** ìˆìŠµë‹ˆë‹¤. ë°œì£¼ë¥¼ ê²€í† í•´ì£¼ì„¸ìš”. (ìƒìœ„ 30ê°œ í‘œì‹œ)")
        st.dataframe(alert, use_container_width=True, height=360)

# ==========================================================
# Render: Dashboard tab (Lot ìª½ ì „ë°˜ í˜„í™©)
# ==========================================================
def render_dashboard():
    b_total = len(binder_df)
    s_total = len(single_df)

    c_b_judge = find_col(binder_df, "íŒì •")
    b_ng = int((binder_df[c_b_judge] == "ë¶€ì í•©").sum()) if c_b_judge and (c_b_judge in binder_df.columns) else 0
    s_ng = int((single_df[c_s_judge] == "ë¶€ì í•©").sum()) if c_s_judge and (c_s_judge in single_df.columns) else 0

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("ë°”ì¸ë” ê¸°ë¡", f"{b_total:,}")
    c2.metric("ë°”ì¸ë” ë¶€ì í•©", f"{b_ng:,}")
    c3.metric("ë‹¨ì¼ìƒ‰ ê¸°ë¡", f"{s_total:,}")
    c4.metric("ë‹¨ì¼ìƒ‰(ì ë„) ë¶€ì í•©", f"{s_ng:,}")

    st.divider()
    st.subheader("ë‹¨ì¼ìƒ‰ ë°ì´í„° ëª©ë¡(í•„í„°)")

    if not (c_s_date and c_s_visc and c_s_pc and (c_s_date in single_df.columns) and (c_s_visc in single_df.columns) and (c_s_pc in single_df.columns)):
        st.warning("ë‹¨ì¼ìƒ‰ ì‹œíŠ¸ì—ì„œ ì…ê³ ì¼/ì ë„ì¸¡ì •ê°’/ì œí’ˆì½”ë“œ ì»¬ëŸ¼ì„ ì°¾ì§€ ëª»í–ˆìŠµë‹ˆë‹¤.")
        return

    df = single_df.copy()
    df[c_s_date] = pd.to_datetime(df[c_s_date], errors="coerce")
    dmin, dmax = safe_date_bounds(df[c_s_date])

    f1, f2, f3, f4 = st.columns([1.2, 1.2, 2.2, 2.4])
    with f1:
        start = st.date_input("ì‹œì‘ì¼", value=max(dmin, dmax - dt.timedelta(days=90)))
    with f2:
        end = st.date_input("ì¢…ë£Œì¼", value=dmax)
    with f3:
        pcs = sorted(df[c_s_pc].dropna().astype(str).unique().tolist())
        sel_pc = st.multiselect("ì œí’ˆì½”ë“œ", pcs, default=[])
    with f4:
        cg = sorted({normalize_color_group(x) for x in df[c_s_cg].dropna().tolist()}) if c_s_cg and c_s_cg in df.columns else []
        sel_cg = st.multiselect("ìƒ‰ìƒê³„ì—´", COLOR_KEYS, default=[])

    if start > end:
        start, end = end, start

    df = df[(df[c_s_date].dt.date >= start) & (df[c_s_date].dt.date <= end)]
    if sel_pc:
        df = df[df[c_s_pc].astype(str).isin(sel_pc)]
    if sel_cg and c_s_cg and c_s_cg in df.columns:
        df = df[df[c_s_cg].apply(normalize_color_group).isin(sel_cg)]

    view = pd.DataFrame({
        "ì…ê³ ì¼": df[c_s_date].dt.date,
        "ìƒ‰ìƒêµ°": df[c_s_cg].apply(normalize_color_group) if c_s_cg and (c_s_cg in df.columns) else None,
        "ì œí’ˆì½”ë“œ": df[c_s_pc].apply(normalize_product_code),
        "ë‹¨ì¼ìƒ‰Lot": df[c_s_lot] if c_s_lot and (c_s_lot in df.columns) else None,
        "ì‚¬ìš©ë°”ì¸ë”Lot": df[c_s_blot] if c_s_blot and (c_s_blot in df.columns) else None,
        "ì ë„(cP)": pd.to_numeric(df[c_s_visc].astype(str).str.replace(",", "", regex=False), errors="coerce"),
        "ì ë„íŒì •": df[c_s_judge] if c_s_judge and (c_s_judge in df.columns) else None,
    }).dropna(subset=["ì…ê³ ì¼"]).sort_values("ì…ê³ ì¼", ascending=False)

    st.dataframe(view, use_container_width=True, height=360)

# ==========================================================
# Render: Binder IO tab (íŒŒì¼ ì—…ë¡œë“œ + êµ¬ê¸€ì‹œíŠ¸)
# ==========================================================
def render_binder_io():
    st.subheader("ë°”ì¸ë” ì…ì¶œê³  ë‚´ì—­ (íŒŒì¼ ì—…ë¡œë“œ / êµ¬ê¸€ì‹œíŠ¸)")
    st.caption("âœ… ë°”ì¸ë” ì…ì¶œê³  ë‚´ì—­ íŒŒì¼(.xlsx)ì„ ì—…ë¡œë“œí•˜ë©´ ì—…ë¡œë“œ ì¦‰ì‹œ í‘œê°€ í‘œì‹œë©ë‹ˆë‹¤. êµ¬ê¸€ì‹œíŠ¸ëŠ” ìƒˆë¡œê³ ì¹¨ ì‹œ ìë™ ë°˜ì˜(ìºì‹œ 60ì´ˆ).")

    # ---- file upload ----
    binder_io_file = st.file_uploader("ë°”ì¸ë” ì…ì¶œê³  ë‚´ì—­ íŒŒì¼ ì—…ë¡œë“œ(.xlsx)", type=["xlsx"], key="binder_io_upload")
    if binder_io_file is not None:
        try:
            io_data = load_binder_io_excel(binder_io_file.getvalue(), binder_io_file.name)
            st.success("ì—…ë¡œë“œ íŒŒì¼ì„ ë¶ˆëŸ¬ì™”ìŠµë‹ˆë‹¤.")
            if "HEMA" in io_data and "Silicone" in io_data:
                c1, c2 = st.columns(2)
                with c1:
                    st.markdown("### HEMA (íŒŒì¼)")
                    st.dataframe(io_data["HEMA"], use_container_width=True, height=420)
                with c2:
                    st.markdown("### Silicone (íŒŒì¼)")
                    st.dataframe(io_data["Silicone"], use_container_width=True, height=420)
            else:
                key = list(io_data.keys())[0]
                st.markdown(f"### {key} (íŒŒì¼)")
                st.dataframe(io_data[key], use_container_width=True, height=520)
        except Exception as e:
            st.error("ì—…ë¡œë“œ íŒŒì¼ì„ ì½ì§€ ëª»í–ˆìŠµë‹ˆë‹¤. (íŒŒì¼ í˜•ì‹/ì‹œíŠ¸ êµ¬ì¡° í™•ì¸)")
            st.exception(e)

    st.divider()

    # ---- google sheets ----
    st.subheader("ë°”ì¸ë” ì…ì¶œê³  (Google Sheets ìë™ ë°˜ì˜)")
    try:
        df_hema = read_gsheet_csv(BINDER_SHEET_ID, BINDER_SHEET_HEMA)
        df_sil = read_gsheet_csv(BINDER_SHEET_ID, BINDER_SHEET_SIL)
    except Exception as e:
        st.error("êµ¬ê¸€ì‹œíŠ¸ì—ì„œ ë°ì´í„°ë¥¼ ëª» ë¶ˆëŸ¬ì™”ìŠµë‹ˆë‹¤. (ì‹œíŠ¸ ê³µìœ /ì›¹ê²Œì‹œ/ì‹œíŠ¸ëª…/ID í™•ì¸)")
        st.exception(e)
        return

    for _df in [df_hema, df_sil]:
        dc = detect_date_col(_df)
        if dc:
            _df["_sort_date"] = pd.to_datetime(_df[dc], errors="coerce")
            _df.sort_values(by="_sort_date", ascending=False, inplace=True)
            _df.drop(columns=["_sort_date"], inplace=True)

    c1, c2 = st.columns(2)
    with c1:
        st.markdown("### HEMA (êµ¬ê¸€ì‹œíŠ¸)")
        st.dataframe(df_hema, use_container_width=True, height=420)
    with c2:
        st.markdown("### Silicone (êµ¬ê¸€ì‹œíŠ¸)")
        st.dataframe(df_sil, use_container_width=True, height=420)

    if st.button("ì§€ê¸ˆ ìµœì‹ ê°’ìœ¼ë¡œ ë‹¤ì‹œ ë¶ˆëŸ¬ì˜¤ê¸°"):
        st.cache_data.clear()
        st.rerun()

# ==========================================================
# Render: Input tab (ì—‘ì…€ì— ì €ì¥)
# ==========================================================
def render_input_tab():
    st.markdown('<div class="section-title">ğŸ“ ì‹ ê·œ ì…ë ¥</div>', unsafe_allow_html=True)
    st.markdown('<div class="section-sub">ì—¬ê¸°ì„œ ì…ë ¥í•œ ê°’ì€ í•´ë‹¹ ì—‘ì…€ ì‹œíŠ¸ì— ë°”ë¡œ ì¶”ê°€(append)ë©ë‹ˆë‹¤.</div>', unsafe_allow_html=True)

    if not Path(xlsx_path).exists():
        st.error("Lot ê´€ë¦¬ ì—‘ì…€ íŒŒì¼ì´ ì—†ì–´ì„œ ì €ì¥í•  ìˆ˜ ì—†ìŠµë‹ˆë‹¤. ì¢Œì¸¡ì—ì„œ ì—…ë¡œë“œí•˜ê±°ë‚˜ ê²½ë¡œë¥¼ ì„¤ì •í•´ ì£¼ì„¸ìš”.")
        return

    st.info("âš ï¸ ì—‘ì…€ì´ PCì—ì„œ ì—´ë ¤ ìˆìœ¼ë©´ ì €ì¥ì´ ì‹¤íŒ¨í•  ìˆ˜ ìˆìŠµë‹ˆë‹¤. ì‹¤íŒ¨ ì‹œ ì—‘ì…€ì„ ë‹«ê³  ë‹¤ì‹œ ì‹œë„í•˜ê±°ë‚˜, ì—…ë¡œë“œ íŒŒì¼ë¡œ ì‹¤í–‰ í›„ 'ë‹¤ìš´ë¡œë“œ'ë¡œ ë°›ëŠ” ë°©ì‹ì´ ì•ˆì „í•©ë‹ˆë‹¤.")

    subt1, subt2, subt3 = st.tabs(["ğŸ§ª ë‹¨ì¼ìƒ‰ ì‰í¬ ì‹ ê·œ ì…ë ¥", "ğŸ§´ ë°”ì¸ë” ì œì¡°/ì…ê³  ì‹ ê·œ ì…ë ¥", "â†©ï¸ ë°”ì¸ë” ì—…ì²´ë°˜í™˜ ì…ë ¥"])

    # ---------- ë‹¨ì¼ìƒ‰ ----------
    with subt1:
        headers = get_sheet_headers(xlsx_path, SHEET_SINGLE)
        if not headers:
            st.error(f"ì‹œíŠ¸/í—¤ë”ë¥¼ ì½ì§€ ëª»í–ˆìŠµë‹ˆë‹¤: {SHEET_SINGLE}")
            return

        existing_lots = []
        if c_s_lot and c_s_lot in single_df.columns:
            existing_lots = single_df[c_s_lot].dropna().astype(str).tolist()

        colA, colB = st.columns([1.6, 1.4])
        with colA:
            st.markdown("#### ì…ë ¥")
            with st.form("form_single"):
                in_date = st.date_input("ì…ê³ ì¼", value=dt.date.today())
                product_code = st.text_input("ì œí’ˆì½”ë“œ", value="")
                color_group = st.selectbox("ìƒ‰ìƒêµ°", COLOR_KEYS, index=COLOR_KEYS.index("OTHER"))
                binder_lot = st.text_input("ì‚¬ìš©ëœ ë°”ì¸ë” Lot", value="")

                auto_lot = st.checkbox("ë‹¨ì¼ìƒ‰ì‰í¬ Lot ìë™ ìƒì„±", value=True)
                if auto_lot:
                    lot_preview = next_lot(existing_lots, in_date, fallback_prefix="PCB")
                    single_lot = st.text_input("ë‹¨ì¼ìƒ‰ì‰í¬ Lot", value=lot_preview)
                else:
                    single_lot = st.text_input("ë‹¨ì¼ìƒ‰ì‰í¬ Lot", value="")

                visc = st.number_input("ì ë„ì¸¡ì •ê°’(cP)", min_value=0, value=0, step=100)
                judge = st.selectbox("ì ë„íŒì •", ["ì í•©", "ë¶€ì í•©", ""], index=2)
                note = st.text_input("ë¹„ê³ (ìˆìœ¼ë©´)", value="")

                submit = st.form_submit_button("âœ… ì €ì¥(ë‹¨ì¼ìƒ‰_ìˆ˜ì…ê²€ì‚¬ì— ì¶”ê°€)", use_container_width=True)

        with colB:
            st.markdown("#### ì €ì¥ë  í˜•íƒœ(ë¯¸ë¦¬ë³´ê¸°)")
            preview = {
                "ì…ê³ ì¼": in_date,
                "ì œí’ˆì½”ë“œ": normalize_product_code(product_code),
                "ìƒ‰ìƒêµ°": normalize_color_group(color_group),
                "ì‚¬ìš©ëœ ë°”ì¸ë” Lot": binder_lot,
                "ë‹¨ì¼ìƒ‰ì‰í¬ Lot": single_lot,
                "ì ë„ì¸¡ì •ê°’(cP)": visc if visc != 0 else None,
                "ì ë„íŒì •": judge if judge else None,
                "ë¹„ê³ ": note if note else None,
            }
            st.dataframe(pd.DataFrame([preview]), use_container_width=True, height=220)

        if submit:
            if not preview["ì…ê³ ì¼"] or not preview["ì œí’ˆì½”ë“œ"] or not preview["ë‹¨ì¼ìƒ‰ì‰í¬ Lot"]:
                st.error("ì…ê³ ì¼ / ì œí’ˆì½”ë“œ / ë‹¨ì¼ìƒ‰ì‰í¬ Lot ëŠ” í•„ìˆ˜ì…ë‹ˆë‹¤.")
            else:
                try:
                    append_row_to_xlsx(xlsx_path, SHEET_SINGLE, preview)
                    st.session_state["_toast_msg"] = "ë‹¨ì¼ìƒ‰_ìˆ˜ì…ê²€ì‚¬ì— ì €ì¥ ì™„ë£Œ"
                    st.cache_data.clear()
                    st.success("ì €ì¥ ì™„ë£Œ âœ…")
                    download_xlsx_button(xlsx_path)
                    st.rerun()
                except PermissionError:
                    st.error("ì €ì¥ ì‹¤íŒ¨: ì—‘ì…€ íŒŒì¼ì´ ì—´ë ¤ ìˆê±°ë‚˜ ì ê²¨ ìˆìŠµë‹ˆë‹¤. ì—‘ì…€ì„ ë‹«ê³  ë‹¤ì‹œ ì‹œë„í•˜ì„¸ìš”.")
                    download_xlsx_button(xlsx_path)
                except Exception as e:
                    st.error("ì €ì¥ ì‹¤íŒ¨")
                    st.exception(e)

    # ---------- ë°”ì¸ë” ì œì¡°/ì…ê³  ----------
    with subt2:
        headers = get_sheet_headers(xlsx_path, SHEET_BINDER)
        if not headers:
            st.error(f"ì‹œíŠ¸/í—¤ë”ë¥¼ ì½ì§€ ëª»í–ˆìŠµë‹ˆë‹¤: {SHEET_BINDER}")
            return

        c_lot_b = find_col(binder_df, "Lot(ìë™)")
        existing_b_lots = []
        if c_lot_b and c_lot_b in binder_df.columns:
            existing_b_lots = binder_df[c_lot_b].dropna().astype(str).tolist()

        st.markdown("#### ì…ë ¥")
        with st.form("form_binder"):
            b_date = st.date_input("ì œì¡°/ì…ê³ ì¼", value=dt.date.today(), key="b_date_in")
            auto_lot_b = st.checkbox("ë°”ì¸ë” Lot(ìë™) ìë™ ìƒì„±", value=True)
            if auto_lot_b:
                b_lot = st.text_input("Lot(ìë™)", value=next_lot(existing_b_lots, b_date, fallback_prefix="PLB"))
            else:
                b_lot = st.text_input("Lot(ìë™)", value="")

            b_judge = st.selectbox("íŒì •", ["ì í•©", "ë¶€ì í•©", ""], index=2, key="b_judge_in")
            b_note = st.text_input("ë¹„ê³ (ìˆìœ¼ë©´)", value="", key="b_note_in")

            with st.expander("ì¶”ê°€ í•­ëª©(ì‹œíŠ¸ì— ì»¬ëŸ¼ì´ ìˆìœ¼ë©´ ê°™ì´ ì €ì¥ë¨)"):
                extras = {}
                candidates = ["ë°”ì¸ë”íƒ€ì…", "ë°”ì¸ë”ëª…", "ì œì¡°ëŸ‰(kg)", "ì ë„(cP)", "ì ë„", "íˆ¬ì…ëŸ‰(kg)", "ë‹´ë‹¹", "ì›ë£Œ Lot", "ì›ë£ŒLot"]
                for name in candidates:
                    if name in headers:
                        if ("kg" in name.lower()) or ("ëŸ‰" in name) or ("íˆ¬ì…" in name):
                            extras[name] = st.number_input(name, min_value=0.0, value=0.0, step=1.0)
                        else:
                            extras[name] = st.text_input(name, value="")

            submit_b = st.form_submit_button("âœ… ì €ì¥(ë°”ì¸ë”_ì œì¡°_ì…ê³ ì— ì¶”ê°€)", use_container_width=True)

        if submit_b:
            row = {
                "ì œì¡°/ì…ê³ ì¼": b_date,
                "Lot(ìë™)": b_lot,
                "íŒì •": b_judge if b_judge else None,
                "ë¹„ê³ ": b_note if b_note else None,
            }
            for k, v in extras.items():
                if isinstance(v, (int, float)) and v == 0:
                    row[k] = None
                else:
                    row[k] = v if str(v).strip() else None

            if not row.get("ì œì¡°/ì…ê³ ì¼") or not row.get("Lot(ìë™)"):
                st.error("ì œì¡°/ì…ê³ ì¼ / Lot(ìë™) ì€ í•„ìˆ˜ì…ë‹ˆë‹¤.")
            else:
                try:
                    append_row_to_xlsx(xlsx_path, SHEET_BINDER, row)
                    st.session_state["_toast_msg"] = "ë°”ì¸ë”_ì œì¡°_ì…ê³ ì— ì €ì¥ ì™„ë£Œ"
                    st.cache_data.clear()
                    st.success("ì €ì¥ ì™„ë£Œ âœ…")
                    download_xlsx_button(xlsx_path)
                    st.rerun()
                except PermissionError:
                    st.error("ì €ì¥ ì‹¤íŒ¨: ì—‘ì…€ íŒŒì¼ì´ ì—´ë ¤ ìˆê±°ë‚˜ ì ê²¨ ìˆìŠµë‹ˆë‹¤. ì—‘ì…€ì„ ë‹«ê³  ë‹¤ì‹œ ì‹œë„í•˜ì„¸ìš”.")
                    download_xlsx_button(xlsx_path)
                except Exception as e:
                    st.error("ì €ì¥ ì‹¤íŒ¨")
                    st.exception(e)

        st.divider()
        st.markdown("#### ìµœê·¼ ë°”ì¸ë” ê¸°ë¡(ìƒìœ„ 30)")
        st.dataframe(binder_df.tail(30).iloc[::-1], use_container_width=True, height=320)

    # ---------- ë°”ì¸ë” ì—…ì²´ë°˜í™˜ ----------
    with subt3:
        headers = get_sheet_headers(xlsx_path, SHEET_BINDER_RETURN)
        if not headers:
            st.error(f"ì‹œíŠ¸/í—¤ë”ë¥¼ ì½ì§€ ëª»í–ˆìŠµë‹ˆë‹¤: {SHEET_BINDER_RETURN}")
            return

        st.markdown("#### ì…ë ¥")
        with st.form("form_return"):
            r_date = st.date_input("ì¼ì", value=dt.date.today(), key="r_date")
            r_type = st.text_input("ë°”ì¸ë”íƒ€ì…", value="")
            r_name = st.text_input("ë°”ì¸ë”ëª…", value="")
            r_lot = st.text_input("ë°”ì¸ë” Lot", value="")
            r_qty = st.number_input("ë°˜í™˜ëŸ‰(kg)", min_value=0.0, value=0.0, step=1.0)
            r_note = st.text_input("ë¹„ê³ ", value="")

            submit_r = st.form_submit_button("âœ… ì €ì¥(ë°”ì¸ë”_ì—…ì²´ë°˜í™˜ì— ì¶”ê°€)", use_container_width=True)

        if submit_r:
            row = {
                "ì¼ì": r_date,
                "ë°”ì¸ë”íƒ€ì…": r_type if r_type else None,
                "ë°”ì¸ë”ëª…": r_name if r_name else None,
                "ë°”ì¸ë” Lot": r_lot if r_lot else None,
                "ë°˜í™˜ëŸ‰(kg)": r_qty if r_qty != 0 else None,
                "ë¹„ê³ ": r_note if r_note else None,
            }
            if not row["ì¼ì"] or not row["ë°”ì¸ë” Lot"] or row["ë°˜í™˜ëŸ‰(kg)"] is None:
                st.error("ì¼ì / ë°”ì¸ë” Lot / ë°˜í™˜ëŸ‰(kg)ì€ í•„ìˆ˜ì…ë‹ˆë‹¤.")
            else:
                try:
                    append_row_to_xlsx(xlsx_path, SHEET_BINDER_RETURN, row)
                    st.session_state["_toast_msg"] = "ë°”ì¸ë”_ì—…ì²´ë°˜í™˜ì— ì €ì¥ ì™„ë£Œ"
                    st.cache_data.clear()
                    st.success("ì €ì¥ ì™„ë£Œ âœ…")
                    download_xlsx_button(xlsx_path)
                    st.rerun()
                except PermissionError:
                    st.error("ì €ì¥ ì‹¤íŒ¨: ì—‘ì…€ íŒŒì¼ì´ ì—´ë ¤ ìˆê±°ë‚˜ ì ê²¨ ìˆìŠµë‹ˆë‹¤. ì—‘ì…€ì„ ë‹«ê³  ë‹¤ì‹œ ì‹œë„í•˜ì„¸ìš”.")
                    download_xlsx_button(xlsx_path)
                except Exception as e:
                    st.error("ì €ì¥ ì‹¤íŒ¨")
                    st.exception(e)

        st.divider()
        st.markdown("#### ìµœê·¼ ë°˜í™˜ ê¸°ë¡(ìƒìœ„ 30)")
        st.dataframe(binder_return_df.tail(30).iloc[::-1], use_container_width=True, height=320)

# ==========================================================
# Render: Search tab
# ==========================================================
def render_search():
    st.subheader("ë¹ ë¥¸ê²€ìƒ‰")
    mode = st.selectbox("ê²€ìƒ‰ ì¢…ë¥˜", ["ë°”ì¸ë” Lot", "ë‹¨ì¼ìƒ‰ Lot", "ì œí’ˆì½”ë“œ"])
    q = st.text_input("ê²€ìƒ‰ì–´", placeholder="ì˜ˆ: PCB20250112-01 / PLB25041501 / PL-835-1 ...")

    s_df = single_df.copy()
    if c_s_date and (c_s_date in s_df.columns):
        s_df[c_s_date] = pd.to_datetime(s_df[c_s_date], errors="coerce")
    b_df = binder_df.copy()
    if c_b_date and (c_b_date in b_df.columns):
        b_df[c_b_date] = pd.to_datetime(b_df[c_b_date], errors="coerce")

    def text_filter(df: pd.DataFrame, cols: list[str], text: str) -> pd.DataFrame:
        if not text:
            return df.iloc[0:0]
        t = str(text).strip()
        if not t:
            return df.iloc[0:0]
        mask = None
        for c in cols:
            if c and c in df.columns:
                m = df[c].astype(str).str.contains(t, case=False, na=False)
                mask = m if mask is None else (mask | m)
        if mask is None:
            return df.iloc[0:0]
        return df[mask]

    if mode == "ë°”ì¸ë” Lot":
        c_bl = find_col(b_df, "Lot(ìë™)")
        hit_b = text_filter(b_df, [c_bl], q)
        st.markdown("#### ë°”ì¸ë”_ì œì¡°_ì…ê³ ")
        st.dataframe(add_excel_row_number(hit_b), use_container_width=True)

        if q and c_s_blot and (c_s_blot in s_df.columns):
            hit_s = s_df[s_df[c_s_blot].astype(str).str.contains(str(q).strip(), case=False, na=False)]
            st.markdown("#### ì—°ê²°ëœ ë‹¨ì¼ìƒ‰_ìˆ˜ì…ê²€ì‚¬ (ì‚¬ìš©ëœ ë°”ì¸ë” Lot)")
            st.dataframe(add_excel_row_number(hit_s), use_container_width=True)

    elif mode == "ë‹¨ì¼ìƒ‰ Lot":
        hit = text_filter(s_df, [c_s_lot], q)
        st.dataframe(add_excel_row_number(hit), use_container_width=True)

    else:  # ì œí’ˆì½”ë“œ
        hit = text_filter(s_df, [c_s_pc], q)
        st.dataframe(add_excel_row_number(hit), use_container_width=True)

# ==========================================================
# Render tabs
# ==========================================================
with tab_dash:
    render_dashboard()

with tab_summary:
    render_summary()

with tab_stock:
    render_stock_tab()

with tab_binder:
    render_binder_io()

with tab_input:
    render_input_tab()

with tab_search:
    render_search()
